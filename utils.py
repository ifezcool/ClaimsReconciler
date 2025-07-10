import pandas as pd
import numpy as np
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


load_dotenv('secrets.env')


def extract_schedule_data(df, schedule_col, amount_col):
    """
    Extract schedule number and amount data from dataframe.

    Args:
        df (pandas.DataFrame): Input dataframe
        schedule_col (str): Name of the column containing schedule numbers
        amount_col (str): Name of the column containing amounts

    Returns:
        pandas.DataFrame: DataFrame with schedule numbers and amounts
    """
    # Create a copy of relevant columns only
    result_df = df[[schedule_col, amount_col]].copy()

    # Rename columns for consistency
    result_df.columns = ["Schedule Number", "Amount"]

    # Convert schedule numbers to string for consistency
    result_df["Schedule Number"] = result_df["Schedule Number"].astype(str)

    # Convert amounts to float, handling potential errors
    result_df["Amount"] = pd.to_numeric(result_df["Amount"], errors="coerce")

    # Drop rows with missing schedule numbers or amounts
    result_df = result_df.dropna()

    # Remove any leading/trailing whitespace from schedule numbers
    result_df["Schedule Number"] = result_df["Schedule Number"].str.strip()

    return result_df

def find_missing_schedules(source_df, target_df):
    """
    Find schedule numbers that are in source_df but not in target_df.

    Args:
        source_df (pandas.DataFrame): Source dataframe with "Schedule Number" column
        target_df (pandas.DataFrame): Target dataframe with "Schedule Number" column

    Returns:
        pandas.DataFrame: DataFrame with schedule numbers missing in target_df
    """
    # Get unique schedule numbers from both dataframes
    source_schedules = set(source_df["Schedule Number"].unique())
    target_schedules = set(target_df["Schedule Number"].unique())

    # Find schedules in source but not in target
    missing_schedules = source_schedules - target_schedules

    # Filter the source dataframe to only include the missing schedules
    missing_df = source_df[source_df["Schedule Number"].isin(missing_schedules)]

    return missing_df

def calculate_schedule_amounts(df):
    """
    Calculate the sum of amounts for each unique schedule number.

    Args:
        df (pandas.DataFrame): DataFrame with "Schedule Number" and "Amount" columns

    Returns:
        pandas.DataFrame: DataFrame with schedule numbers and their total amounts
    """
    # Group by schedule number and sum the amounts
    schedule_amounts = df.groupby("Schedule Number")["Amount"].sum().reset_index()

    return schedule_amounts

def generate_reconciliation_report(claims_amounts, finance_amounts):
    """
    Generate a reconciliation report comparing Claims and Finance amounts.

    Args:
        claims_amounts (pandas.DataFrame): DataFrame with Claims schedule numbers and amounts
        finance_amounts (pandas.DataFrame): DataFrame with Finance schedule numbers and amounts

    Returns:
        pandas.DataFrame: Reconciliation report with schedule numbers and amounts from both sources
    """
    # Merge the two dataframes on schedule number
    merged = pd.merge(
        claims_amounts, 
        finance_amounts, 
        on="Schedule Number", 
        how="outer",
        suffixes=(" Claims", " Finance")
    )

    # Rename columns for clarity
    merged.columns = ["Schedule Number", "Claims Amount", "Finance Amount"]

    # Sort by schedule number
    merged = merged.sort_values("Schedule Number")

    # Calculate difference
    merged["Difference"] = merged["Claims Amount"] - merged["Finance Amount"]

    return merged

def generate_enhanced_claims_excel(claims_df, schedule_col, amount_col):
    """
    Generate an Excel file with all claims data plus 4 formula columns.

    Args:
        claims_df (pandas.DataFrame): The original claims dataframe
        schedule_col (str): Name of the schedule column
        amount_col (str): Name of the amount column

    Returns:
        bytes: Excel file as bytes
    """
    import io
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    print(f"🔍 DEBUG UTILS: Starting generate_enhanced_claims_excel")
    print(f"🔍 DEBUG UTILS: Input DF shape: {claims_df.shape}")
    print(f"🔍 DEBUG UTILS: Input DF columns: {list(claims_df.columns)}")
    print(f"🔍 DEBUG UTILS: Schedule col: {schedule_col}")
    print(f"🔍 DEBUG UTILS: Amount col: {amount_col}")

    # Create a copy of the claims dataframe
    enhanced_df = claims_df.copy()
    print(f"🔍 DEBUG UTILS: Created copy of dataframe")

    # Add all the additional columns in the correct order
    additional_columns = [
        'ReviewedDate', 'PostedDate', 'PaidDate',
        'ClaimBatch', 'ClaimNoFnx', 'ClaimNo', 'Correct_ClaimNo',
        'Benefits', 'ProviderClass', 'OpdIpd'
    ]

    for col in additional_columns:
        enhanced_df[col] = ""

    print(f"🔍 DEBUG UTILS: Added {len(additional_columns)} additional columns. New shape: {enhanced_df.shape}")

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Enhanced Claims Data"
    print(f"🔍 DEBUG UTILS: Created workbook and worksheet")

    # Write the dataframe to the worksheet
    print(f"🔍 DEBUG UTILS: Writing dataframe to worksheet...")
    for i, r in enumerate(dataframe_to_rows(enhanced_df, index=False, header=True)):
        ws.append(r)
        if i % 100 == 0:  # Log every 100 rows
            print(f"🔍 DEBUG UTILS: Wrote row {i}")
    print(f"🔍 DEBUG UTILS: Finished writing {len(enhanced_df)} rows to worksheet")

    # Find the column indices for the key columns in the actual dataframe
    # Map common column name variations to standard names
    column_mapping = {
        'PROVIDER_CODE': ['PROVIDER CODE', 'PROVIDER_CODE', 'Provider Code', 'Provider_Code', 'ProviderCode'],
        'ENCOUNTER_DATE_DD_MM_YYYY': ['ENCOUNTER DATE (DD/MM/YYYY)', 'ENCOUNTER_DATE_DD_MM_YYYY', 'ENCOUNTER_DATE', 'Encounter Date', 'Encounter_Date', 'ENC_DATE'],
        'DATE_CLAIM_RECEIVED': ['DATE CLAIM RECEIVED ', 'DATE_CLAIM_RECEIVED', 'Date Claim Received', 'Date_Claim_Received', 'CLAIM_RECEIVED_DATE'],
        'ENROLLEE_NAME': ['ENROLLEE NAME', 'ENROLLEE_NAME', 'Enrollee Name', 'Enrollee_Name', 'EnrolleeName'],
        'MEMBER_NO': ['MEMBER NO', 'MEMBER_NO', 'Member No', 'Member_No', 'MemberNo', 'Member Number']
    }

    # Function to find column by multiple possible names
    def find_column_index(df, possible_names):
        for name in possible_names:
            if name in df.columns:
                return df.columns.get_loc(name) + 1  # +1 for Excel 1-based indexing
        return None

    # Find the actual column indices
    provider_code_idx = find_column_index(enhanced_df, column_mapping['PROVIDER_CODE'])
    encounter_date_idx = find_column_index(enhanced_df, column_mapping['ENCOUNTER_DATE_DD_MM_YYYY'])
    claim_received_idx = find_column_index(enhanced_df, column_mapping['DATE_CLAIM_RECEIVED'])
    enrollee_name_idx = find_column_index(enhanced_df, column_mapping['ENROLLEE_NAME'])
    member_no_idx = find_column_index(enhanced_df, column_mapping['MEMBER_NO'])

    print(f"🔍 DEBUG UTILS: Column indices found:")
    print(f"  provider_code_idx: {provider_code_idx}")
    print(f"  encounter_date_idx: {encounter_date_idx}")
    print(f"  claim_received_idx: {claim_received_idx}")
    print(f"  enrollee_name_idx: {enrollee_name_idx}")
    print(f"  member_no_idx: {member_no_idx}")

    # Find the formula column indices
    claimbatch_idx = enhanced_df.columns.get_loc('ClaimBatch') + 1
    claimnofnx_idx = enhanced_df.columns.get_loc('ClaimNoFnx') + 1
    claimno_idx = enhanced_df.columns.get_loc('ClaimNo') + 1
    correct_claimno_idx = enhanced_df.columns.get_loc('Correct_ClaimNo') + 1

    print(f"🔍 DEBUG UTILS: Formula column indices:")
    print(f"  claimbatch_idx: {claimbatch_idx}")
    print(f"  claimnofnx_idx: {claimnofnx_idx}")
    print(f"  claimno_idx: {claimno_idx}")
    print(f"  correct_claimno_idx: {correct_claimno_idx}")

    # Convert indices to Excel column letters
    if provider_code_idx:
        provider_code_col = get_column_letter(provider_code_idx)
    if encounter_date_idx:
        encounter_date_col = get_column_letter(encounter_date_idx)
    if claim_received_idx:
        claim_received_col = get_column_letter(claim_received_idx)
    if enrollee_name_idx:
        enrollee_name_col = get_column_letter(enrollee_name_idx)
    if member_no_idx:
        member_no_col = get_column_letter(member_no_idx)

    claimbatch_col = get_column_letter(claimbatch_idx)
    claimnofnx_col = get_column_letter(claimnofnx_idx)
    claimno_col = get_column_letter(claimno_idx)
    correct_claimno_col = get_column_letter(correct_claimno_idx)

    # Add formulas for each data row (starting from row 2, since row 1 is headers)
    print(f"🔍 DEBUG UTILS: Adding formulas for {len(enhanced_df)} rows...")
    for row_idx in range(2, len(enhanced_df) + 2):
        if row_idx % 100 == 2:  # Log every 100 rows starting from the first data row
            print(f"🔍 DEBUG UTILS: Adding formulas for row {row_idx}")

        # ClaimBatch formula - month from claim_received (U), year from encounter_date (R)
        if provider_code_idx and encounter_date_idx and claim_received_idx:
            claimbatch_formula = f'=IF({provider_code_col}{row_idx}="NIL","9999999"&(IF(LEN(MONTH({claim_received_col}{row_idx}))=1,"0"&MONTH({claim_received_col}{row_idx}),MONTH({claim_received_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2),{provider_code_col}{row_idx}&(IF(LEN(MONTH({claim_received_col}{row_idx}))=1,"0"&MONTH({claim_received_col}{row_idx}),MONTH({claim_received_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2))'
            ws[f'{claimbatch_col}{row_idx}'] = claimbatch_formula

        # ClaimNoFnx formula - simplified logic as per user specification
        if enrollee_name_idx:
            if row_idx == 2:  # First data row
                claimnofnx_formula = f'=COUNTIF(${enrollee_name_col}$2:{enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF(${enrollee_name_col}$2:{enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1'
            else:
                claimnofnx_formula = f'=IF({enrollee_name_col}{row_idx}={enrollee_name_col}{row_idx-1},{claimnofnx_col}{row_idx-1}+1,COUNTIF(${enrollee_name_col}$2:{enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF(${enrollee_name_col}$2:{enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1)'
            ws[f'{claimnofnx_col}{row_idx}'] = claimnofnx_formula

        # ClaimNo formula - month from claim_received (U), year from encounter_date (R)
        if member_no_idx and encounter_date_idx and claim_received_idx and enrollee_name_idx:
            if row_idx == 2:  # First data row
                claimno_formula = f'={member_no_col}{row_idx}&(IF(LEN(MONTH({claim_received_col}{row_idx}))=1,"0"&MONTH({claim_received_col}{row_idx}),MONTH({claim_received_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2)&(COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1)'
            else:
                claimno_formula = f'={member_no_col}{row_idx}&(IF(LEN(MONTH({claim_received_col}{row_idx}))=1,"0"&MONTH({claim_received_col}{row_idx}),MONTH({claim_received_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2)&IF({enrollee_name_col}{row_idx}={enrollee_name_col}{row_idx-1},{claimnofnx_col}{row_idx-1}+1,COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1)'
            ws[f'{claimno_col}{row_idx}'] = claimno_formula

        # Correct_ClaimNo formula - all dates from encounter_date (R)
        if member_no_idx and encounter_date_idx and enrollee_name_idx:
            if row_idx == 2:  # First data row
                correct_claimno_formula = f'={member_no_col}{row_idx}&(IF(LEN(DAY({encounter_date_col}{row_idx}))=1,"0"&DAY({encounter_date_col}{row_idx}),DAY({encounter_date_col}{row_idx})))&(IF(LEN(MONTH({encounter_date_col}{row_idx}))=1,"0"&MONTH({encounter_date_col}{row_idx}),MONTH({encounter_date_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2)&(COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1)'
            else:
                correct_claimno_formula = f'={member_no_col}{row_idx}&(IF(LEN(DAY({encounter_date_col}{row_idx}))=1,"0"&DAY({encounter_date_col}{row_idx}),DAY({encounter_date_col}{row_idx})))&(IF(LEN(MONTH({encounter_date_col}{row_idx}))=1,"0"&MONTH({encounter_date_col}{row_idx}),MONTH({encounter_date_col}{row_idx})))&RIGHT(YEAR({encounter_date_col}{row_idx}),2)&IF({enrollee_name_col}{row_idx}={enrollee_name_col}{row_idx-1},{claimnofnx_col}{row_idx-1}+1,COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})-COUNTIF({enrollee_name_col}$2:${enrollee_name_col}{row_idx},{enrollee_name_col}{row_idx})+1)'
            ws[f'{correct_claimno_col}{row_idx}'] = correct_claimno_formula

    # Save to BytesIO
    print(f"🔍 DEBUG UTILS: Saving workbook to BytesIO...")
    excel_output = io.BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    file_size = len(excel_output.getvalue())
    print(f"🔍 DEBUG UTILS: Excel file created successfully. Size: {file_size} bytes")

    return excel_output.getvalue()

def send_variance_email(variance_type, missing_schedules=None, amount_variances=None, date_errors=None):
    """
    Send email notification for variances found during reconciliation.

    Args:
        variance_type (str): Type of variance - "missing_schedules" or "amount_variances"
        missing_schedules (list): List of schedule numbers missing in finance
        amount_variances (list): List of dictionaries with variance details
    """
    # Email configuration
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    sender_email = os.getenv("OFFICE_SENDER_EMAIL")
    sender_password = os.getenv("OUTLOOK_APP_PASSWORD")
    recipient_email = "ifeoluwa.adeniyi@avonhealthcare.com"
    cc_email = ["ifeoluwa.adeniyi@avonhealthcare.com"
                "adedamola.ayeni@avonhealthcare.com",
                "adebola.adesoyin@avonhealthcare.com",
                "claims_officers@avonhealthcare.com",
                "bi_dataanalytics@avonhealthcare.com",
                "financedepartment@avonhealthcare.com"
                ]
    #financedepartment@avonhealthcare.com
    # Check if credentials are available
    if not sender_email or not sender_password:
        print("❌ Gmail credentials not found in environment variables")
        print("Please set GMAIL_SENDER_EMAIL and GMAIL_APP_PASSWORD in Secrets")
        return

    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        #Choose the correct CC list depending on variance/error type
        if variance_type == "date_validation_errors":
            cc_list = ["ifeoluwa.adeniyi@avonhealthcare.com",#"claims_officers@avonhealthcare.com",
                       "claims_officers@avonhealthcare.com",
                        "bi_dataanalytics@avonhealthcare.com"]
        else:
            cc_list = cc_email
        msg['Cc'] = ", ".join(cc_list)
        msg['Subject'] = f"Claims Reconciliation Alert - {variance_type.replace('_', ' ').title()}"

        # Create email body based on variance type
        if variance_type == "missing_schedules" and missing_schedules:
            body = f"""
Dear Finance,

This is an automated notification from the Claims Reconciliation System.

CRITICAL ALERT: The following Schedule Numbers (SCH NO) were found in the Claims department report but are MISSING in the Finance department report:

Missing Schedule Numbers:
{chr(10).join([f"- {sch}" for sch in missing_schedules])}

Total Missing Schedules: {len(missing_schedules)}

This indicates that these schedules were sent by Claims but have not been received or processed by Finance. Please investigate and take appropriate action.

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Best regards,
Claims Reconciliation System
Avon Healthcare Ltd.
"""

        elif variance_type == "amount_variances" and amount_variances:
            body = f"""
Dear Claims/Finance Dept,

This is an automated notification from the Claims Reconciliation System.

AMOUNT VARIANCE ALERT: The following Schedule Numbers (SCH NO) have DIFFERENT AMOUNTS between Claims and Finance departments:

Amount Variances:
"""
            for variance in amount_variances:
                body += f"""
- SCH NO: {variance['schedule']}
  Claims Amount: {variance['claims_amount']:,.2f}
  Finance Amount: {variance['finance_amount']:,.2f}
  Difference: {variance['difference']:,.2f}
"""

            body += f"""
Total Schedules with Amount Variances: {len(amount_variances)}

Please review these discrepancies and ensure the amounts are properly reconciled.Edit the live sheet on the sharepoint.

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Best regards,
Claims Reconciliation System
Avon Healthcare Ltd.
"""

        elif variance_type == "date_validation_errors" and date_errors:
            body = f"""
Dear Claims,

This is an automated notification from the Claims Reconciliation System.

DATE VALIDATION ERROR ALERT: The following Schedule Numbers (SCH NO) have ENCOUNTER DATES that are AFTER the DATE CLAIM RECEIVED, which indicates a data entry error:

Date Validation Errors:
"""
            # Count occurrences of each unique error
            error_counts = {}
            for error in date_errors:
                error_key = (error['schedule'], error['encounter_date'], error['claim_received_date'])
                if error_key in error_counts:
                    error_counts[error_key] += 1
                else:
                    error_counts[error_key] = 1
            
            for (schedule, encounter_date, claim_received_date), count in error_counts.items():
                body += f"""
- SCH NO: {schedule}
  Encounter Date: {encounter_date}
  Date Claim Received: {claim_received_date}
  Issue: Encounter date cannot be after the claim was received
  Number of Records: {count}
"""

            body += f"""


Please review these data entry errors and edit the live sheet on the sharepoint to correct the dates. The encounter date should not be after the date the claim was received.

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Best regards,
Claims Reconciliation System
Avon Healthcare Ltd.
"""

        else:
            return  # No valid variance data provided

        # Attach body to email
        msg.attach(MIMEText(body, 'plain'))

        # Create SMTP session
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable security
        server.login(sender_email, sender_password)

        allrecipients = [recipient_email] + cc_list

        # Send email
        text = msg.as_string()
        server.sendmail(sender_email, allrecipients, text)
        server.quit()

        print(f"📧 Email sent successfully for {variance_type}")

    except Exception as e:
        print(f"❌ Failed to send email: {str(e)}")
        raise e