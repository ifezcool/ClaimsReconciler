import streamlit as st
import pandas as pd
import numpy as np
import io
import plotly.express as px
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import base64
from session_manager import save_upload, get_session_data, get_available_sessions
from utils import (
    extract_schedule_data,
    find_missing_schedules,
    calculate_schedule_amounts,
    generate_reconciliation_report,
    generate_enhanced_claims_excel,
    send_variance_email
)

st.set_page_config(
    page_title="Claims Reconciliation Tool",
    page_icon="📊",
    layout="wide"
)

# Initialize session state variables if they don't exist
if 'uploaded_claims_file' not in st.session_state:
    st.session_state.uploaded_claims_file = None
if 'uploaded_finance_file' not in st.session_state:
    st.session_state.uploaded_finance_file = None
if 'claims_sheet' not in st.session_state:
    st.session_state.claims_sheet = None
if 'finance_sheet' not in st.session_state:
    st.session_state.finance_sheet = None
if 'claims_schedule_col' not in st.session_state:
    st.session_state.claims_schedule_col = None
if 'claims_amount_col' not in st.session_state:
    st.session_state.claims_amount_col = None
if 'finance_schedule_col' not in st.session_state:
    st.session_state.finance_schedule_col = None
if 'finance_amount_col' not in st.session_state:
    st.session_state.finance_amount_col = None
if 'generate_claims_excel' not in st.session_state:
    st.session_state.generate_claims_excel = False

# Page navigation
page = st.sidebar.selectbox(
    "Select Page",
    ["Claims Reconciliation", "Appeals Compilation","DB_Upload"]
)

if page == "Appeals Compilation":
    from appeals_page import show_appeals_page
    show_appeals_page()
    st.stop()
elif page == "DB_Upload":
    from DB_Upload import render_dbpage
    render_dbpage()
    st.stop()
elif page == "AppealsUpload":
    from AppealsUpload import render_appeals_upload
    render_appeals_upload()
    st.stop()

st.title("Claims Reconciliation Tool")
st.markdown("""
This application automates the reconciliation process between Claims and Finance department reports.
Upload the Excel reports from both departments to identify discrepancies in schedule numbers and amounts.
""")

# Initialize session state if not already initialized
if 'department' not in st.session_state:
    st.session_state.department = None
if 'session_loaded' not in st.session_state:
    st.session_state.session_loaded = False

# Define file change callbacks
def on_claims_file_change():
    if 'claims_file_uploader' in st.session_state and st.session_state.claims_file_uploader is not None:
        # Make a copy to ensure persistence
        file_copy = io.BytesIO()
        st.session_state.claims_file_uploader.seek(0)
        file_copy.write(st.session_state.claims_file_uploader.read())
        file_copy.seek(0)
        st.session_state.uploaded_claims_file = file_copy

def on_finance_file_change():
    if 'finance_file_uploader' in st.session_state and st.session_state.finance_file_uploader is not None:
        # Make a copy to ensure persistence
        file_copy = io.BytesIO()
        st.session_state.finance_file_uploader.seek(0)
        file_copy.write(st.session_state.finance_file_uploader.read())
        file_copy.seek(0)
        st.session_state.uploaded_finance_file = file_copy

# Department selection and session management
st.header("Department Selection")

# Role selection
dept_col1, dept_col2 = st.columns(2)
with dept_col1:
    department_options = ["Select Department", "Claims Department", "Finance Department", "Reconciliation Manager"]
    selected_department = st.selectbox(
        "Select your department:",
        options=department_options,
        index=0
    )

    # Handle department change while preserving uploaded files
    if selected_department != "Select Department" and selected_department != st.session_state.department:
        # Store any currently loaded files in session state before switching departments
        if 'claims_file_uploader' in st.session_state and st.session_state.claims_file_uploader is not None:
            on_claims_file_change()

        if 'finance_file_uploader' in st.session_state and st.session_state.finance_file_uploader is not None:
            on_finance_file_change()

        # Update department
        st.session_state.department = selected_department.split()[0].lower()  # Store 'claims', 'finance', or 'reconciliation'

# Get available sessions
available_sessions = get_available_sessions()
with dept_col2:
    if selected_department == "Reconciliation Manager" and available_sessions:
        session_options = ["Current Week"] + available_sessions
        selected_session = st.selectbox(
            "Select session to load:",
            options=session_options,
            index=0
        )

        if st.button("Load Selected Session"):
            session_id = None if selected_session == "Current Week" else selected_session
            session_data = get_session_data(session_id)

            # Store session data in session state
            st.session_state.selected_session = session_id
            st.session_state.session_data = session_data
            st.session_state.session_loaded = True

            # Display status messages for both departments
            col1, col2 = st.columns(2)
            with col1:
                if session_data['claims'] is not None:
                    claims_time = datetime.fromisoformat(session_data['claims']['timestamp'])
                    st.success(f"✅ Claims file from {claims_time.strftime('%Y-%m-%d %H:%M')}")
                else:
                    st.warning("⚠️ No Claims file in this session")

            with col2:
                if session_data['finance'] is not None:
                    finance_time = datetime.fromisoformat(session_data['finance']['timestamp'])
                    st.success(f"✅ Finance file from {finance_time.strftime('%Y-%m-%d %H:%M')}")
                else:
                    st.warning("⚠️ No Finance file in this session")

            # Overall status
            if session_data['claims'] is not None and session_data['finance'] is not None:
                st.success("Both Claims and Finance files loaded successfully!")
            elif session_data['claims'] is None and session_data['finance'] is None:
                st.error("No data available in the selected session.")

# File upload section
st.header("Upload Files")

# Determine which uploads to show based on department
if st.session_state.department == 'claims':
    # Only Claims upload
    st.markdown("### Claims Department File")
    claims_file = st.file_uploader(
        "Upload Claims Department Excel Report (PAYMENT SCHEDULE)",
        type=["xlsx"],
        key="claims_file_uploader",
        on_change=on_claims_file_change
    )
    # Retain finance file from session state
    finance_file = st.session_state.uploaded_finance_file
elif st.session_state.department == 'finance':
    # Only Finance upload
    st.markdown("### Finance Department File")
    finance_file = st.file_uploader(
        "Upload Finance Department Excel Report (Finance claims reconciliation)",
        type=["xlsx"],
        key="finance_file_uploader",
        on_change=on_finance_file_change
    )
    # Retain claims file from session state
    claims_file = st.session_state.uploaded_claims_file
else:
    # Both uploads for reconciliation view
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### Claims Department File")
        claims_file = st.file_uploader(
            "Upload Claims Department Excel Report (PAYMENT SCHEDULE)",
            type=["xlsx"],
            key="claims_file_uploader",
            on_change=on_claims_file_change
        )

    with col2:
        st.markdown("### Finance Department File")
        finance_file = st.file_uploader(
            "Upload Finance Department Excel Report (Finance claims reconciliation)",
            type=["xlsx"],
            key="finance_file_uploader",
            on_change=on_finance_file_change
        )

# Use session state files if available
if claims_file is None and st.session_state.uploaded_claims_file is not None:
    claims_file = st.session_state.uploaded_claims_file
    # Need to seek to beginning as the file might have been read already
    claims_file.seek(0)

if finance_file is None and st.session_state.uploaded_finance_file is not None:
    finance_file = st.session_state.uploaded_finance_file
    # Need to seek to beginning as the file might have been read already
    finance_file.seek(0)

# Always load the current session data for reconciliation manager
if st.session_state.department == 'reconciliation':
    # Get the latest session data (always reload this for reconciliation manager)
    latest_session_data = get_session_data()

    # If in reconciliation view, always try to load both files from the session
    if not claims_file and 'claims' in latest_session_data and latest_session_data['claims'] is not None:
        claims_data_copy = latest_session_data['claims']['file_data']
        claims_file = io.BytesIO()
        claims_data_copy.seek(0)
        claims_file.write(claims_data_copy.read())
        claims_file.seek(0)
        st.session_state.claims_sheet = latest_session_data['claims']['sheet_name']
        st.session_state.claims_schedule_col = latest_session_data['claims']['schedule_col']
        st.session_state.claims_amount_col = latest_session_data['claims']['amount_col']

    if not finance_file and 'finance' in latest_session_data and latest_session_data['finance'] is not None:
        finance_data_copy = latest_session_data['finance']['file_data']
        finance_file = io.BytesIO()
        finance_data_copy.seek(0)
        finance_file.write(finance_data_copy.read())
        finance_file.seek(0)
        st.session_state.finance_sheet = latest_session_data['finance']['sheet_name']
        st.session_state.finance_schedule_col = latest_session_data['finance']['schedule_col']
        st.session_state.finance_amount_col = latest_session_data['finance']['amount_col']

# Process files when both are uploaded
if claims_file and finance_file:
    try:
        st.header("File Analysis")

        # Load files with pandas
        claims_xls = pd.ExcelFile(claims_file)
        finance_xls = pd.ExcelFile(finance_file)

        # Display sheets in each file
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("### Claims File Sheets")
            claims_sheets = claims_xls.sheet_names
            selected_claims_sheet = st.selectbox(
                "Select the sheet with payment schedules:", 
                claims_sheets,
                index=0
            )

        with col2:
            st.markdown("### Finance File Sheets")
            finance_sheets = finance_xls.sheet_names
            selected_finance_sheet = st.selectbox(
                "Select the sheet with claims received:", 
                finance_sheets,
                index=0 if "Claims recei" in finance_sheets[0] else 0
            )

        # Load selected sheets
        with st.spinner("Loading and processing data..."):
            claims_df = pd.read_excel(claims_file, sheet_name=selected_claims_sheet)
            finance_df = pd.read_excel(finance_file, sheet_name=selected_finance_sheet)

            # Display preview of loaded data
            st.subheader("Data Preview")
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Claims Department Data")
                st.dataframe(claims_df.head())

            with col2:
                st.markdown("#### Finance Department Data")
                st.dataframe(finance_df.head())

            # Extract schedule data
            claims_schedule_col, claims_amount_col = None, None
            finance_schedule_col, finance_amount_col = None, None

            # Detect column names for schedule numbers
            possible_schedule_cols = ["SCH NO", "Claim Batch No/Sch No", "Schedule No", "Schedule Number", "SCH_NO"]
            possible_amount_cols_claims = ["HOD RECOMMD. AMOUNT", "HOD AMOUNT", "RECOMMENDED AMOUNT", "AMOUNT"]
            possible_amount_cols_finance = ["Claims_Advised_Amount", "Advised_Amount", "Claim Amount", "AMOUNT"]

            # Find appropriate columns in claims dataframe
            for col in possible_schedule_cols:
                if col in claims_df.columns:
                    claims_schedule_col = col
                    break

            for col in possible_amount_cols_claims:
                if col in claims_df.columns:
                    claims_amount_col = col
                    break

            # Find appropriate columns in finance dataframe
            for col in possible_schedule_cols:
                if col in finance_df.columns:
                    finance_schedule_col = col
                    break

            for col in possible_amount_cols_finance:
                if col in finance_df.columns:
                    finance_amount_col = col
                    break

            # Column selection form
            st.subheader("Column Selection")
            st.info("Please confirm the columns that contain schedule numbers and amounts in both files.")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Claims Department Columns")
                claims_schedule_col = st.selectbox(
                    "Schedule Number Column (Claims):", 
                    claims_df.columns,
                    index=claims_df.columns.get_loc(claims_schedule_col) if claims_schedule_col else 0
                )
                claims_amount_col = st.selectbox(
                    "Amount Column (Claims):", 
                    claims_df.columns,
                    index=claims_df.columns.get_loc(claims_amount_col) if claims_amount_col else 0
                )

            with col2:
                st.markdown("#### Finance Department Columns")
                finance_schedule_col = st.selectbox(
                    "Schedule Number Column (Finance):", 
                    finance_df.columns,
                    index=finance_df.columns.get_loc(finance_schedule_col) if finance_schedule_col else 0
                )
                finance_amount_col = st.selectbox(
                    "Amount Column (Finance):", 
                    finance_df.columns,
                    index=finance_df.columns.get_loc(finance_amount_col) if finance_amount_col else 0
                )

            # Handle single department uploads - Save to session first
            dept = st.session_state.department
            if dept in ['claims', 'finance'] and st.button(f"Upload {dept.capitalize()} Data"):
                # Get current file and settings
                if st.session_state.department == 'claims' and claims_file:
                    # Save claims file to session
                    file_copy = io.BytesIO()
                    claims_file.seek(0)
                    file_copy.write(claims_file.read())
                    file_copy.seek(0)

                    try:
                        save_upload(
                            'claims', 
                            file_copy, 
                            selected_claims_sheet, 
                            claims_schedule_col, 
                            claims_amount_col
                        )
                        st.success(f"Claims data uploaded successfully! Finance department will be notified to upload their data.")
                    except Exception as e:
                        st.error(f"Error saving upload: {str(e)}")

                elif st.session_state.department == 'finance' and finance_file:
                    # Save finance file to session
                    file_copy = io.BytesIO()
                    finance_file.seek(0)
                    file_copy.write(finance_file.read())
                    file_copy.seek(0)

                    try:
                        save_upload(
                            'finance', 
                            file_copy, 
                            selected_finance_sheet, 
                            finance_schedule_col, 
                            finance_amount_col
                        )
                        st.success(f"Finance data uploaded successfully! The reconciliation manager will be notified.")
                    except Exception as e:
                        st.error(f"Error saving upload: {str(e)}")

            # Status for Reconciliation Manager
            if st.session_state.department == 'reconciliation':
                st.subheader("Uploaded Files Status")
                status_col1, status_col2 = st.columns(2)

                with status_col1:
                    if claims_file:
                        st.success("✅ Claims Department file loaded successfully")
                    else:
                        st.warning("⚠️ Claims Department file not available")

                with status_col2:
                    if finance_file:
                        st.success("✅ Finance Department file loaded successfully")
                    else:
                        st.warning("⚠️ Finance Department file not available")

                if not claims_file or not finance_file:
                    st.info("You need both Claims and Finance files to perform reconciliation. Either upload them directly or load from a previous session.")

                # Button to process the reconciliation (for manager only)
                if st.button("Process Reconciliation"):
                    with st.spinner("Processing reconciliation..."):
                        # Extract and process data
                        claims_data = extract_schedule_data(claims_df, claims_schedule_col, claims_amount_col)
                        finance_data = extract_schedule_data(finance_df, finance_schedule_col, finance_amount_col)

                        # Calculate aggregated amounts for each schedule
                        claims_amounts = calculate_schedule_amounts(claims_data)
                        finance_amounts = calculate_schedule_amounts(finance_data)

                    # Store variables in session state
                    st.session_state.claims_data = claims_data
                    st.session_state.finance_data = finance_data
                    st.session_state.claims_amounts = claims_amounts
                    st.session_state.finance_amounts = finance_amounts
                    st.session_state.reconciliation_processed = True
                    st.session_state.emails_sent = False  # Reset email flag for new reconciliation

            # Check if reconciliation data is available
            if ('claims_amounts' in st.session_state and 'finance_amounts' in st.session_state and 
                st.session_state.get('reconciliation_processed', False)):

                # Define tolerance for floating-point precision issues
                tolerance = 0.01

                # Find missing schedules
                missing_in_finance = find_missing_schedules(st.session_state.claims_data, st.session_state.finance_data)
                missing_in_claims = find_missing_schedules(st.session_state.finance_data, st.session_state.claims_data)

                # Generate reconciliation report
                reconciliation_report = generate_reconciliation_report(
                    st.session_state.claims_amounts, st.session_state.finance_amounts
                )

                # Display results
                st.header("Reconciliation Results")

                # Send emails only once per reconciliation (manually triggered)
                if st.button("Send Notification Emails") and not st.session_state.get('emails_sent', False):
                    # Check for missing schedules and send email
                    if not missing_in_finance.empty:
                        try:
                            send_variance_email(
                                variance_type="missing_schedules",
                                missing_schedules=missing_in_finance["Schedule Number"].tolist(),
                                amount_variances=None
                            )
                            st.info("📧 Email notification sent for missing schedules")
                        except Exception as e:
                            st.warning(f"Failed to send email notification: {str(e)}")

                    # Check for date validation errors (encounter date after claim received date)
                    date_validation_errors = []
                    
                    # Find encounter date and claim received date columns
                    encounter_date_col = None
                    claim_received_col = None
                    
                    encounter_date_variations = ['ENCOUNTER DATE (DD/MM/YYYY)', 'ENCOUNTER_DATE_DD_MM_YYYY', 'ENCOUNTER_DATE', 'Encounter Date', 'Encounter_Date', 'ENC_DATE']
                    claim_received_variations = ['DATE CLAIM RECEIVED ', 'DATE_CLAIM_RECEIVED', 'Date Claim Received', 'Date_Claim_Received', 'CLAIM_RECEIVED_DATE']
                    
                    for col in encounter_date_variations:
                        if col in claims_df.columns:
                            encounter_date_col = col
                            break
                    
                    for col in claim_received_variations:
                        if col in claims_df.columns:
                            claim_received_col = col
                            break
                    
                    if encounter_date_col and claim_received_col:
                        # Check each row for date validation issues
                        for idx, row in claims_df.iterrows():
                            try:
                                encounter_date = pd.to_datetime(row[encounter_date_col], errors='coerce')
                                claim_received_date = pd.to_datetime(row[claim_received_col], errors='coerce')
                                
                                if pd.notna(encounter_date) and pd.notna(claim_received_date):
                                    if encounter_date > claim_received_date:
                                        sch_no = str(row[claims_schedule_col]) if claims_schedule_col in row else "Unknown"
                                        date_validation_errors.append({
                                            'schedule': sch_no,
                                            'encounter_date': encounter_date.strftime('%d/%m/%Y'),
                                            'claim_received_date': claim_received_date.strftime('%d/%m/%Y')
                                        })
                            except Exception:
                                # Skip rows with invalid date formats
                                continue
                    
                    # Send email for date validation errors
                    if date_validation_errors:
                        try:
                            send_variance_email(
                                variance_type="date_validation_errors",
                                missing_schedules=None,
                                amount_variances=None,
                                date_errors=date_validation_errors
                            )
                            st.warning(f"📧 Email sent for {len(date_validation_errors)} date validation errors")
                        except Exception as e:
                            st.warning(f"Failed to send email notification for date errors: {str(e)}")

                    # Check for amount variances and send email
                    common_schedules = set(st.session_state.claims_amounts["Schedule Number"]).intersection(
                        set(st.session_state.finance_amounts["Schedule Number"])
                    )
                    amount_mismatch = reconciliation_report[
                        (~reconciliation_report.isna().any(axis=1)) & 
                        (abs(reconciliation_report['Claims Amount'] - reconciliation_report['Finance Amount']) > tolerance)
                    ]
                    
                    if not amount_mismatch.empty:
                        # Send email for amount variances
                        try:
                            variance_details = []
                            for _, row in amount_mismatch.iterrows():
                                variance_details.append({
                                    'schedule': row['Schedule Number'],
                                    'claims_amount': row['Claims Amount'],
                                    'finance_amount': row['Finance Amount'],
                                    'difference': row['Difference']
                                })
                            
                            send_variance_email(
                                variance_type="amount_variances",
                                missing_schedules=None,
                                amount_variances=variance_details
                            )
                            st.info("📧 Email notification sent for amount variances")
                        except Exception as e:
                            st.warning(f"Failed to send email notification: {str(e)}")

                    # Mark emails as sent
                    st.session_state.emails_sent = True

                # Missing Schedules
                st.subheader("Claims Schedules Missing in Finance (Critical)")
                if missing_in_finance.empty:
                    st.success("No schedules missing in Finance - All schedules sent by Claims were received by Finance")
                else:
                    st.error(f"{len(missing_in_finance)} schedules sent by Claims but not found in Finance")
                    st.dataframe(missing_in_finance.style.highlight_max(axis=0, color='red'), use_container_width=True)

                    # Add total amount for missing schedules
                    total_missing_amount = missing_in_finance["Amount"].sum()
                    st.error(f"Total amount missing: {total_missing_amount:,.2f}")

                # Reconciliation Report
                st.subheader("Amount Reconciliation")
                st.markdown("This table shows the comparison of amounts for each schedule number between Claims and Finance.")

                # Highlight differences in the report (with tolerance for floating-point precision)
                def highlight_diff(row):
                    if pd.isna(row['Claims Amount']) or pd.isna(row['Finance Amount']):
                        return ['background-color: yellow'] * len(row)
                    elif abs(row['Claims Amount'] - row['Finance Amount']) > tolerance:
                        return ['background-color: lightcoral'] * len(row)
                    else:
                        return [''] * len(row)

                formatted_report = reconciliation_report.style.apply(highlight_diff, axis=1)

                # Calculate discrepancy metrics based on Claims perspective
                total_claims_schedules = len(st.session_state.claims_amounts)

                # Schedules that are in both Claims and Finance
                common_schedules = set(st.session_state.claims_amounts["Schedule Number"]).intersection(
                    set(st.session_state.finance_amounts["Schedule Number"])
                )

                # Matching amounts (with tolerance for floating-point precision)
                matching_schedules_df = reconciliation_report.dropna(subset=['Claims Amount', 'Finance Amount'])
                matching_schedules = sum(abs(matching_schedules_df['Claims Amount'] - matching_schedules_df['Finance Amount']) <= tolerance)

                # Discrepancies
                discrepancies = total_claims_schedules - matching_schedules

                # Missing values and amount mismatches for other calculations
                missing_values = reconciliation_report[reconciliation_report.isna().any(axis=1)]
                amount_mismatch = reconciliation_report[
                    (~reconciliation_report.isna().any(axis=1)) & 
                    (abs(reconciliation_report['Claims Amount'] - reconciliation_report['Finance Amount']) > tolerance)
                ]

                # Display metrics
                col1, col2, col3 = st.columns(3)
                col1.metric("Total Claims Schedules", total_claims_schedules)
                col2.metric("Matching Amounts", matching_schedules)
                col3.metric("Discrepancies", discrepancies)

                # Calculate financial metrics
                total_claims_amount = st.session_state.claims_amounts["Amount"].sum()

                # Get only the matching schedules for finance amount
                matching_schedules_only = reconciliation_report.dropna(subset=['Claims Amount', 'Finance Amount'])
                matching_finance_amount = matching_schedules_only['Finance Amount'].sum()

                # Calculate variance
                total_variance = total_claims_amount - matching_finance_amount

                # Display financial metrics
                st.subheader("Financial Summary")
                col1, col2, col3 = st.columns(3)
                col1.metric("Total Claims Amount", f"{total_claims_amount:,.2f}")
                col2.metric("Matching Finance Amount", f"{matching_finance_amount:,.2f}",
                           help="Total amount from Finance where schedules match with Claims")
                col3.metric("Total Variance", f"{total_variance:,.2f}",
                           delta=f"{(total_variance/total_claims_amount*100 if total_claims_amount else 0):.2f}%")

                # Show detailed report
                st.dataframe(formatted_report)

                # Visualization
                st.subheader("Visual Comparison")

                # Filter out missing values for the visualization
                vis_data = reconciliation_report.dropna()

                if not vis_data.empty:
                    # Scatter plot comparing amounts
                    fig = px.scatter(
                        vis_data, 
                        x="Claims Amount", 
                        y="Finance Amount",
                        hover_data=["Schedule Number"],
                        labels={
                            "Claims Amount": "Amount Sent by Claims",
                            "Finance Amount": "Amount Received by Finance"
                        },
                        title="Claims vs Finance Amounts"
                    )

                    # Add a diagonal line representing perfect match
                    max_val = max(vis_data["Claims Amount"].max(), vis_data["Finance Amount"].max())
                    fig.add_scatter(
                        x=[0, max_val], 
                        y=[0, max_val], 
                        mode='lines', 
                        line=dict(color='green', dash='dash'),
                        name="Perfect Match"
                    )

                    st.plotly_chart(fig, use_container_width=True)

                    # Bar chart for discrepancies
                    if not amount_mismatch.empty:
                        # Create comparative bar chart
                        amount_mismatch_melted = pd.melt(
                            amount_mismatch.reset_index(), 
                            id_vars=["Schedule Number"],
                            value_vars=["Claims Amount", "Finance Amount"],
                            var_name="Department", 
                            value_name="Amount"
                        )

                        fig2 = px.bar(
                            amount_mismatch_melted,
                            x="Schedule Number",
                            y="Amount",
                            color="Department",
                            barmode="group",
                            title="Schedules with Amount Discrepancies",
                            labels={"Schedule Number": "Schedule Number", "Amount": "Amount"}
                        )
                        st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.info("No data available for visualization after removing missing values.")

                # Download section
                st.header("Download Results")

                # Enhanced Claims Excel button
                if st.button("Generate Claims Data with Formula Columns"):
                    st.session_state.generate_claims_excel = True
                    st.rerun()

                # Check if claims excel should be generated
                if st.session_state.generate_claims_excel:
                    with st.spinner("Generating Claims Excel with formula columns..."):
                        try:
                            # Generate the enhanced claims Excel
                            enhanced_claims_excel = generate_enhanced_claims_excel(claims_df, claims_schedule_col, claims_amount_col)

                            filename = f"Enhanced Claims Data with Formulas {pd.Timestamp.now().strftime('%d %b %Y')}.xlsx"

                            st.download_button(
                                label="📊 Download Enhanced Claims Excel",
                                data=enhanced_claims_excel,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.success("Enhanced Claims Excel generated successfully!")

                            # Reset the flag
                            st.session_state.generate_claims_excel = False
                        except Exception as e:
                            st.error(f"Error generating enhanced claims Excel: {str(e)}")
                            st.session_state.generate_claims_excel = False

                # Prepare data for Excel report
                common_schedules = set(st.session_state.claims_amounts["Schedule Number"]).intersection(
                    set(st.session_state.finance_amounts["Schedule Number"])
                )

                # Filter the reconciliation report to include only common schedules
                matching_report = reconciliation_report[
                    reconciliation_report["Schedule Number"].isin(common_schedules)
                ].copy()

                # Create a formatted reconciliation report
                formatted_recon_df = pd.DataFrame({
                    "Claim Batch No/Sch No": matching_report["Schedule Number"],
                    "Year": pd.Timestamp.now().year,
                    "Week Period": f"{pd.Timestamp.now().strftime('%d %b')} - {(pd.Timestamp.now() + pd.Timedelta(days=6)).strftime('%d %b')}",
                    "Processing Platform": "MANUAL",
                    "Claim Type": "FRESH CLAIM",
                    "Claims_Advised_Amount": matching_report["Claims Amount"],
                    "Finance_Recognized_Amount": matching_report["Finance Amount"],
                    "Variance": matching_report["Difference"].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else ""),
                    "Comments": "",
                    "Status": "",
                })

                # For amount mismatches among common schedules
                if not amount_mismatch.empty:
                    mismatch_schedules = amount_mismatch["Schedule Number"].tolist()
                    formatted_recon_df.loc[formatted_recon_df["Claim Batch No/Sch No"].isin(mismatch_schedules), "Status"] = "Amount Mismatch"
                    formatted_recon_df.loc[formatted_recon_df["Claim Batch No/Sch No"].isin(mismatch_schedules), "Comments"] = "Variance in reported amounts"

                # Sort by schedule number
                formatted_recon_df = formatted_recon_df.sort_values("Claim Batch No/Sch No")

                # Generate Excel file for download
                excel_output = io.BytesIO()
                with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
                    # Write to Excel - main reconciliation report sheet
                    workbook = writer.book
                    worksheet = workbook.create_sheet('Claims Reconciliation Report')
                    title_cell = worksheet.cell(row=1, column=1, value='BI Unit - Claims Received Weekly Report')
                    title_cell.font = openpyxl.styles.Font(bold=True)

                    # Add all the data starting from row 2
                    for r_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(formatted_recon_df, index=False, header=True), 2):
                        for c_idx, value in enumerate(row, 1):
                            worksheet.cell(row=r_idx, column=c_idx, value=value)

                    # Also include the detailed sheets for reference
                    if not missing_in_finance.empty:
                        missing_in_finance.to_excel(writer, sheet_name='Missing in Finance', index=False)
                    if not missing_in_claims.empty:
                        missing_in_claims.to_excel(writer, sheet_name='Missing in Claims', index=False)
                    if not amount_mismatch.empty:
                        amount_mismatch.to_excel(writer, sheet_name='Amount Discrepancies', index=False)

                excel_output.seek(0)

                # Excel download
                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_output,
                    file_name=f"BI Unit - Claims Reconciliation Weekly Report {pd.Timestamp.now().strftime('%d %b %Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Summary
                st.header("Reconciliation Summary")

                # Create a DataFrame for the pie chart
                labels = ['Matching Amounts', 'Missing in Finance', 'Amount Mismatches']
                values = [
                    matching_schedules, 
                    len(missing_in_finance), 
                    discrepancies - len(missing_in_finance)
                ]
                values = [max(0, v) for v in values]

                if sum(values) > 0:
                    fig = px.pie(
                        names=labels, 
                        values=values,
                        title="Distribution of Claims Reconciliation Issues",
                        color_discrete_sequence=px.colors.qualitative.Safe
                    )
                    st.plotly_chart(fig, use_container_width=True)

                st.markdown(f"""
                ### Key Metrics:
                - **Total Claims Schedules**: {total_claims_schedules}
                - **Schedules with Matching Amounts**: {matching_schedules} ({matching_schedules/total_claims_schedules*100:.1f}% of Claims total)
                - **Total Discrepancies**: {discrepancies} ({discrepancies/total_claims_schedules*100:.1f}% of Claims total)

                ### Critical Issues:
                - **Schedules Missing in Finance**: {len(missing_in_finance)} 
                  - Total amount missing: {missing_in_finance["Amount"].sum():,.2f}

                ### Financial Analysis:
                - **Total Claims Amount**: {total_claims_amount:,.2f}
                - **Matching Finance Amount**: {matching_finance_amount:,.2f} (only from schedules present in both)
                - **Total Amount Variance**: {total_variance:,.2f} ({(total_variance/total_claims_amount*100 if total_claims_amount else 0):.2f}% of Claims total)

                ### Other Issues:
                - **Amount Mismatches in Common Schedules**: {len(amount_mismatch)}
                """)

    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
        st.error("Please ensure you've selected the correct sheets and columns.")
else:
    st.info("Please upload both Claims and Finance department Excel files to begin reconciliation.")

# Instructions section
with st.expander("How to Use This Tool"):
    st.markdown("""
    ### Instructions

    1. **Upload Files**:
       - Upload the Claims Department excel file (usually titled 'PAYMENT SCHEDULE SENT TO FINANCE')
       - Upload the Finance Department excel file (usually titled 'Finance claims reconciliation report')

    2. **Select Sheets**:
       - From the dropdown menus, select the correct sheets that contain the data to reconcile

    3. **Confirm Columns**:
       - Verify that the automatically detected columns are correct, or select the appropriate columns
       - For Claims: You need the schedule number column (usually 'SCH NO') and amount column (usually 'HOD RECOMMD. AMOUNT')
       - For Finance: You need the schedule number column (usually 'Claim Batch No/Sch No') and amount column (usually 'Claims_Advised_Amount')

    4. **Process Reconciliation**:
       - Click the "Process Reconciliation" button to analyze the data

    5. **Review Results**:
       - Check for missing schedules in either department
       - Examine the amount discrepancies between departments
       - Use the visualizations to identify patterns in discrepancies

    6. **Download Reports**:
       - Download the detailed reconciliation report in Excel format for record keeping or further analysis
       - Download a formatted PDF report with all key findings and statistics for easy sharing
    """)

# Footer
st.markdown("---")
st.markdown("Claims Reconciliation Tool | For internal use only")
